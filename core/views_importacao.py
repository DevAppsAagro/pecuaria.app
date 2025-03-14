from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
from django.db.models import Q
import pandas as pd
from decimal import Decimal
from datetime import datetime
from .models import Animal, Lote, CategoriaAnimal, Raca, Pasto, Fazenda
import json
import traceback
from django.http import JsonResponse
import tempfile
import os
import numpy as np

def criar_planilha_modelo(raca=None, categoria=None, lote=None, pasto=None, quantidade=10):
    """
    Cria uma planilha modelo para importação de animais
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Animais"
    
    # Definir cores e estilos
    titulo_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    
    # Título e instruções (sem mesclagem excessiva)
    ws.cell(row=1, column=1, value="PLANILHA DE IMPORTAÇÃO DE ANIMAIS").font = Font(bold=True, size=14)
    
    # Adicionar instruções sem mesclagem excessiva
    ws.cell(row=2, column=1, value="INSTRUÇÕES:").font = Font(bold=True)
    ws.cell(row=3, column=1, value="1. Preencha os campos obrigatórios (marcados com *)").font = Font(italic=True)
    ws.cell(row=4, column=1, value="2. Não altere o nome das colunas").font = Font(italic=True)
    ws.cell(row=5, column=1, value="3. Datas devem estar no formato DD/MM/AAAA").font = Font(italic=True)
    
    # Cabeçalhos - linha 7
    headers = [
        "Brinco Visual*", "Brinco Eletrônico", 
        "Nome do Lote*", "Data de Nascimento* (DD/MM/AAAA)", "Data de Entrada* (DD/MM/AAAA)",
        "Raça*", "Categoria*", "Peso de Entrada (kg)*", "Valor de Compra (R$)*",
        "Pasto Atual*"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        # Ajustar largura da coluna
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 2, 15)
    
    # Buscar valores para os campos se forem fornecidos
    raca_nome = ''
    categoria_nome = ''
    lote_nome = ''
    pasto_nome = ''
    
    # Importações dentro da função para evitar problemas de circular import
    from django.apps import apps
    Raca = apps.get_model('core', 'Raca')
    CategoriaAnimal = apps.get_model('core', 'CategoriaAnimal')
    Lote = apps.get_model('core', 'Lote')
    Pasto = apps.get_model('core', 'Pasto')
    
    if raca:
        try:
            raca_obj = Raca.objects.get(id=raca)
            raca_nome = raca_obj.nome
        except:
            pass
            
    if categoria:
        try:
            categoria_obj = CategoriaAnimal.objects.get(id=categoria)
            categoria_nome = categoria_obj.nome
        except:
            pass
            
    if lote:
        try:
            lote_obj = Lote.objects.get(id=lote)
            lote_nome = lote_obj.id_lote
        except:
            pass
            
    if pasto:
        try:
            pasto_obj = Pasto.objects.get(id=pasto)
            pasto_nome = pasto_obj.id_pasto
        except:
            pass
    
    # Preencher linhas de exemplo
    for i in range(1, quantidade + 1):
        row = 7 + i
        ws.cell(row=row, column=1, value=f"BR{i:05d}")  # Brinco Visual
        # Brinco Eletrônico fica em branco
        
        if lote_nome:
            ws.cell(row=row, column=3, value=lote_nome)  # Lote
        # Data de Nascimento fica em branco
        # Data de Entrada fica em branco
        
        if raca_nome:
            ws.cell(row=row, column=6, value=raca_nome)  # Raça
        if categoria_nome:
            ws.cell(row=row, column=7, value=categoria_nome)  # Categoria
        # Peso fica em branco
        # Valor fica em branco
        
        if pasto_nome:
            ws.cell(row=row, column=10, value=pasto_nome)  # Pasto
    
    return wb

def download_planilha_modelo(request):
    """
    View para download da planilha modelo de importação de animais
    """
    # Obter parâmetros do formulário
    raca_id = request.GET.get('raca')
    categoria_id = request.GET.get('categoria')
    lote_id = request.GET.get('lote')
    pasto_id = request.GET.get('pasto')
    quantidade = request.GET.get('quantidade', 10)
    
    try:
        quantidade = int(quantidade)
        if quantidade < 1:
            quantidade = 10
        elif quantidade > 100:  # Limitar a um máximo razoável
            quantidade = 100
    except (ValueError, TypeError):
        quantidade = 10
    
    # Criar a planilha com os parâmetros personalizados
    wb = criar_planilha_modelo(
        raca=raca_id, 
        categoria=categoria_id, 
        lote=lote_id, 
        pasto=pasto_id, 
        quantidade=quantidade
    )
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=importacao_animais_personalizada.xlsx'
    wb.save(response)
    return response

@login_required
def import_animais(request):
    """Função para importar animais a partir de uma planilha Excel."""
    context = {
        'active_page': 'importacao_animais',
        'fazendas': Fazenda.objects.filter(usuario=request.user),
    }
    
    if request.method != 'POST' or not request.FILES.get('arquivo'):
        return render(request, 'animais/animal_import.html', context)
    
    try:
        # Criar log de depuração detalhado
        print("\n==== INICIANDO IMPORTAÇÃO DE ANIMAIS ====")
        arquivo = request.FILES['arquivo']
        print(f"Arquivo: {arquivo.name}, Tamanho: {arquivo.size} bytes")
        
        # Salvar cópia do arquivo para depuração
        import tempfile
        import os
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"import_debug_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        with open(temp_file, 'wb') as f:
            for chunk in arquivo.chunks():
                f.write(chunk)
        print(f"Cópia de debug salva em: {temp_file}")
        
        # === LEITURA SIMPLIFICADA DA PLANILHA ===
        
        # Tentar abrir o arquivo como Excel e obter informações básicas
        arquivo.seek(0)
        try:
            xls = pd.ExcelFile(arquivo)
            print(f"Planilhas disponíveis: {xls.sheet_names}")
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Erro ao abrir o arquivo: {str(e)}. Verifique se é um arquivo Excel válido.'
            })
        
        # Tentar ler a planilha pulando linhas até encontrar conteúdo válido
        arquivo.seek(0)
        dfs = []
        
        # Testar várias configurações de cabeçalho e escolher a melhor
        for header_row in range(0, 10):
            try:
                df = pd.read_excel(arquivo, header=header_row)
                arquivo.seek(0)
                
                # Verificar se tem pelo menos uma coluna
                if len(df.columns) >= 5:  # Esperamos pelo menos 5 colunas importantes
                    dfs.append((header_row, df))
            except:
                pass
        
        if not dfs:
            return JsonResponse({
                'status': 'error',
                'message': 'Não foi possível ler o conteúdo da planilha. Verifique se o formato está correto.'
            })
        
        # === ANÁLISE E SELEÇÃO DO MELHOR DATAFRAME ===
        
        # Funções auxiliares para normalizar textos e comparar similaridade
        def normalize_text(text):
            if not isinstance(text, str):
                text = str(text)
            return text.lower().strip()
        
        def text_similarity(text1, text2):
            text1 = normalize_text(text1)
            text2 = normalize_text(text2)
            
            # Correspondência exata
            if text1 == text2:
                return 1.0
                
            # Verificar se um contém o outro
            if text1 in text2 or text2 in text1:
                return 0.8
                
            # Verificar palavras comuns
            words1 = set(text1.split())
            words2 = set(text2.split())
            common = words1.intersection(words2)
            
            if common:
                return len(common) / max(len(words1), len(words2))
                
            return 0.0
        
        # Colunas esperadas e seus possíveis nomes alternativos
        expected_columns = {
            'Brinco Visual': ['brinco', 'visual', 'id', 'identificação', 'identificacao', 'numero', 'código', 'codigo', 'tag', 
                           'ident', 'identif', 'brinc', 'animal', 'identificador', 'cod', 'num', 'número', 'n°', 'nº', 'no', 'n.',
                           'registro', 'identificação do animal', 'identificacao do animal', 'codigo do animal', 'código do animal'],
            'Nome do Lote': ['lote', 'grupo', 'nome lote', 'nomelote', 'nome_lote', 'lote nome', 'id_lote', 'id lote', 'nome do grupo',
                          'id do lote', 'identificação do lote', 'nome lote', 'lotes', 'grupo de animais', 'conjunto'],
            'Raça': ['raça', 'raca', 'breed', 'tipo de animal', 'genetica', 'genética', 'origem', 'tipo', 'especie', 'espécie'],
            'Categoria': ['categoria', 'tipo', 'classe', 'classificação', 'classificacao', 'group', 'grupo', 'tipo animal', 
                       'tipo de animal', 'finalidade', 'funcao', 'função'],
            'Pasto Atual': ['pasto', 'paddock', 'area', 'área', 'local', 'localização', 'localizacao', 'location', 
                          'pasto atual', 'pasto corrente', 'pastagem', 'campo', 'curral', 'piquete']
        }
        
        # Analisar cada dataframe para encontrar o que tem mais colunas esperadas
        best_df = None
        best_mapping = {}
        best_score = 0
        best_header = 0
        
        # Log para ajudar na depuração
        print("\n===== TENTANDO IDENTIFICAR COLUNAS =====")
        
        for header_row, df in dfs:
            print(f"\nTestando cabeçalho na linha {header_row}")
            print(f"Colunas encontradas: {list(df.columns)}")
            
            # Normalizar nomes das colunas
            normalized_columns = {}
            for col in df.columns:
                normalized_columns[normalize_text(col)] = col
            
            # Tentar mapear colunas esperadas
            mapping = {}
            score = 0
            
            for expected_col, alternatives in expected_columns.items():
                # Log para depuração
                print(f"\nProcurando coluna '{expected_col}':")
                
                # Primeiro tentar correspondência exata
                exact_match = normalize_text(expected_col)
                if exact_match in normalized_columns:
                    mapping[expected_col] = normalized_columns[exact_match]
                    score += 1
                    print(f"  ✓ Correspondência exata: '{normalized_columns[exact_match]}'")
                    continue
                
                # Se não encontrou, tentar as alternativas
                matched = False
                for alt in alternatives:
                    alt_normalized = normalize_text(alt)
                    if alt_normalized in normalized_columns:
                        mapping[expected_col] = normalized_columns[alt_normalized]
                        score += 1
                        print(f"  ✓ Correspondência alternativa: '{alt}' → '{normalized_columns[alt_normalized]}'")
                        matched = True
                        break
                
                if matched:
                    continue
                
                # Se ainda não encontrou, procurar por similaridade
                best_match = None
                best_sim = 0.3  # Limite mínimo de similaridade
                best_match_description = ""
                
                # Primeiro verificar se alguma coluna contém o nome esperado
                for col_orig in df.columns:
                    col_norm = normalize_text(col_orig)
                    
                    # Pular colunas já mapeadas
                    if col_orig in mapping.values():
                        continue
                    
                    # Verificar se o nome da coluna contém o termo esperado
                    for term in [expected_col] + alternatives:
                        term_norm = normalize_text(term)
                        # Correspondência exata (case insensitive)
                        if col_norm == term_norm:
                            best_match = col_orig
                            best_sim = 1.0
                            best_match_description = f"correspondência exata (case insensitive): '{col_orig}'"
                            break
                        # Coluna contém o termo
                        elif term_norm in col_norm:
                            if best_sim < 0.9:
                                best_match = col_orig
                                best_sim = 0.9
                                best_match_description = f"coluna contém o termo '{term}': '{col_orig}'"
                        # Termo contém a coluna
                        elif col_norm in term_norm:
                            if best_sim < 0.8:
                                best_match = col_orig
                                best_sim = 0.8
                                best_match_description = f"termo '{term}' contém a coluna: '{col_orig}'"
                    
                    # Verificar similaridade com o nome esperado
                    sim = text_similarity(expected_col, col_norm)
                    if sim > best_sim:
                        best_sim = sim
                        best_match = col_orig
                        best_match_description = f"similaridade de texto ({sim:.2f}): '{col_orig}'"
                    
                    # Verificar similaridade com alternativas
                    for alt in alternatives:
                        sim = text_similarity(alt, col_norm)
                        if sim > best_sim:
                            best_sim = sim
                            best_match = col_orig
                            best_match_description = f"similaridade com alternativa '{alt}' ({sim:.2f}): '{col_orig}'"
                
                if best_match:
                    mapping[expected_col] = best_match
                    score += best_sim  # Adicionar pontuação parcial
                    print(f"  ✓ Encontrado por {best_match_description}")
                else:
                    print(f"  ✗ Não encontrado")
            
            # Verificar se este dataframe é melhor que o anterior
            print(f"\nMapeamento encontrado: {mapping}")
            print(f"Pontuação: {score:.2f} (quanto maior, melhor)")
            
            if score > best_score:
                best_score = score
                best_df = df
                best_mapping = mapping
                best_header = header_row
                print(f"Este é o melhor mapeamento até agora!")
        
        # Se não conseguiu encontrar todas as colunas por nome, tentar por posição
        if best_df is None or len(best_mapping) < len(expected_columns):
            print("\n\n===== TENTANDO FALLBACK POR POSIÇÃO =====")
            
            # Determinar quais colunas estão faltando
            missing = [col for col in expected_columns.keys() if col not in best_mapping]
            
            # Se temos um DataFrame, vamos tentar posições fixas para as colunas que faltam
            if best_df is not None and len(best_df.columns) >= 5:
                # Posições típicas para colunas de importação de animais
                # Estas posições são baseadas em formatos comuns de planilhas
                position_mapping = {
                    'Brinco Visual': 0,  # Geralmente a primeira coluna
                    'Brinco Eletrônico': 1,  # Geralmente a segunda coluna
                    'Nome do Lote': 2,  # Geralmente a terceira coluna
                    'Raça': 3,  # Geralmente a quarta coluna
                    'Categoria': 4,  # Geralmente a quinta coluna
                    'Pasto Atual': 5,  # Geralmente a sexta coluna
                    'Data de Nascimento': 6,  # Geralmente a sétima coluna
                    'Data de Entrada': 7,  # Geralmente a oitava coluna
                    'Peso de Entrada': 8,  # Geralmente a nona coluna
                    'Valor de Compra': 9,  # Geralmente a décima coluna
                }
                
                # Adicionar colunas faltantes usando posições
                for col in missing:
                    pos = position_mapping.get(col)
                    if pos is not None and pos < len(best_df.columns):
                        best_mapping[col] = best_df.columns[pos]
                        print(f"Adicionado '{col}' por posição: {best_df.columns[pos]} (coluna {pos})")
            
            # Se ainda não temos um DataFrame bom, vamos usar o primeiro que encontramos
            if best_df is None and dfs:
                header_row, best_df = dfs[0]
                best_header = header_row
                
                # Tentar mapear usando posições
                if len(best_df.columns) >= 5:
                    best_mapping = {}
                    position_mapping = {
                        'Brinco Visual': 0,
                        'Nome do Lote': 1,
                        'Raça': 2,
                        'Categoria': 3,
                        'Pasto Atual': 4,
                    }
                    
                    for col, pos in position_mapping.items():
                        if pos < len(best_df.columns):
                            best_mapping[col] = best_df.columns[pos]
                            print(f"Mapeando '{col}' para coluna {pos}: {best_df.columns[pos]}")
        
        # Verificação final
        if best_df is None or len(best_mapping) < len(expected_columns):
            # Determinar quais colunas ainda estão faltando
            missing = [col for col in expected_columns.keys() if col not in best_mapping]
            
            # Se tudo falhar, tente usar a primeira linha como dados e nomear colunas manualmente
            if len(dfs) > 0:
                print("\n\n===== ÚLTIMA TENTATIVA: NOMEANDO COLUNAS MANUALMENTE =====")
                header_row, df_raw = dfs[0]
                
                # Criar nomes de colunas genéricos
                col_names = ['col' + str(i) for i in range(len(df_raw.columns))]
                df_raw.columns = col_names
                
                # Mapear colunas obrigatórias por posição
                force_mapping = {
                    'Brinco Visual': 'col0',
                    'Nome do Lote': 'col1',
                    'Raça': 'col2',
                    'Categoria': 'col3',
                    'Pasto Atual': 'col4',
                }
                
                # Usar este mapeamento se for melhor que o anterior
                if len(force_mapping) > len(best_mapping):
                    best_df = df_raw
                    best_mapping = force_mapping
                    best_header = header_row
                    print("Aplicado mapeamento forçado por posição.")
            
            # Se ainda estiver faltando colunas
            missing = [col for col in expected_columns.keys() if col not in best_mapping]
            if missing:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Não foi possível identificar todas as colunas necessárias. Colunas faltando: {", ".join(missing)}. Por favor, verifique se a planilha possui estas colunas com nomes semelhantes.'
                })
        
        # === PROCESSAMENTO DAS LINHAS ===
        
        print(f"Usando planilha com cabeçalho na linha {best_header}")
        print(f"Mapeamento de colunas encontrado: {best_mapping}")
        
        # Mostrar primeiras linhas para debug
        print("\nPrimeiras 5 linhas da planilha:")
        print(best_df.head(5).to_string())
        
        # Processar cada linha da planilha
        imported_count = 0
        errors = []
        
        for index, row in best_df.iterrows():
            try:
                print(f"\nProcessando linha {index+1}")
                
                # ==== BRINCO VISUAL ====
                brinco_col = best_mapping['Brinco Visual']
                brinco_raw = row[brinco_col]
                
                # Verificar se está vazio
                if pd.isna(brinco_raw):
                    errors.append(f"Linha {index+1}: Brinco visual está vazio")
                    continue
                
                # Conversão para string - tratamento específico para cada tipo
                print(f"Brinco bruto: {brinco_raw} (tipo: {type(brinco_raw).__name__})")
                
                if isinstance(brinco_raw, (int, float, np.number)):
                    # Números - converte para inteiro para remover casas decimais
                    brinco_visual = str(int(brinco_raw))
                elif isinstance(brinco_raw, (datetime, pd.Timestamp)):
                    # Datas - converte para string no formato YYYY-MM-DD
                    brinco_visual = str(brinco_raw.date())
                else:
                    # Strings e outros tipos
                    brinco_visual = str(brinco_raw).strip()
                
                print(f"Brinco processado: '{brinco_visual}'")
                
                # Verificar se ficou vazio após processamento
                if not brinco_visual:
                    errors.append(f"Linha {index+1}: Brinco visual está vazio após processamento")
                    continue
                
                # Verificar duplicidade
                if Animal.objects.filter(brinco_visual=brinco_visual).exists():
                    errors.append(f"Linha {index+1}: Já existe um animal com o brinco visual '{brinco_visual}'")
                    continue
                
                # ==== LOTE ====
                lote_col = best_mapping['Nome do Lote']
                lote_raw = row[lote_col]
                
                if pd.isna(lote_raw):
                    errors.append(f"Linha {index+1}: Nome do lote está vazio")
                    continue
                
                # Converter para string
                if isinstance(lote_raw, (int, float, np.number)):
                    lote_nome = str(int(lote_raw))
                elif isinstance(lote_raw, (datetime, pd.Timestamp)):
                    lote_nome = str(lote_raw.date())
                else:
                    lote_nome = str(lote_raw).strip()
                
                # Buscar o lote no banco de dados
                try:
                    lote = Lote.objects.get(nome=lote_nome, usuario=request.user)
                except Lote.DoesNotExist:
                    errors.append(f"Linha {index+1}: Lote '{lote_nome}' não encontrado")
                    continue
                
                # ==== RAÇA ====
                raca_col = best_mapping['Raça']
                raca_raw = row[raca_col]
                
                if pd.isna(raca_raw):
                    errors.append(f"Linha {index+1}: Raça está vazia")
                    continue
                
                # Converter para string
                if isinstance(raca_raw, (int, float, np.number)):
                    raca_nome = str(int(raca_raw))
                elif isinstance(raca_raw, (datetime, pd.Timestamp)):
                    raca_nome = str(raca_raw.date())
                else:
                    raca_nome = str(raca_raw).strip()
                
                # Buscar a raça no banco de dados
                try:
                    raca = Raca.objects.filter(Q(usuario=request.user) | Q(usuario__isnull=True)).get(nome=raca_nome)
                except Raca.DoesNotExist:
                    errors.append(f"Linha {index+1}: Raça '{raca_nome}' não encontrada")
                    continue
                
                # ==== CATEGORIA ====
                categoria_col = best_mapping['Categoria']
                categoria_raw = row[categoria_col]
                
                if pd.isna(categoria_raw):
                    errors.append(f"Linha {index+1}: Categoria está vazia")
                    continue
                
                # Converter para string
                if isinstance(categoria_raw, (int, float, np.number)):
                    categoria_nome = str(int(categoria_raw))
                elif isinstance(categoria_raw, (datetime, pd.Timestamp)):
                    categoria_nome = str(categoria_raw.date())
                else:
                    categoria_nome = str(categoria_raw).strip()
                
                # Buscar a categoria no banco de dados
                try:
                    categoria = CategoriaAnimal.objects.filter(Q(usuario=request.user) | Q(usuario__isnull=True)).get(nome=categoria_nome)
                except CategoriaAnimal.DoesNotExist:
                    errors.append(f"Linha {index+1}: Categoria '{categoria_nome}' não encontrada")
                    continue
                
                # ==== PASTO ====
                pasto_col = best_mapping['Pasto Atual']
                pasto_raw = row[pasto_col]
                
                if pd.isna(pasto_raw):
                    errors.append(f"Linha {index+1}: Pasto está vazio")
                    continue
                
                # Converter para string
                if isinstance(pasto_raw, (int, float, np.number)):
                    pasto_nome = str(int(pasto_raw))
                elif isinstance(pasto_raw, (datetime, pd.Timestamp)):
                    pasto_nome = str(pasto_raw.date())
                else:
                    pasto_nome = str(pasto_raw).strip()
                
                # Buscar o pasto no banco de dados
                try:
                    pasto = Pasto.objects.get(nome=pasto_nome, fazenda=lote.fazenda)
                except Pasto.DoesNotExist:
                    errors.append(f"Linha {index+1}: Pasto '{pasto_nome}' não encontrado")
                    continue
                
                # ==== CAMPOS OPCIONAIS ====
                
                # Brinco Eletrônico (opcional)
                brinco_eletronico = None
                if 'Brinco Eletrônico' in best_mapping:
                    brinco_e_col = best_mapping['Brinco Eletrônico']
                    brinco_e_raw = row[brinco_e_col]
                    
                    if not pd.isna(brinco_e_raw):
                        if isinstance(brinco_e_raw, (int, float, np.number)):
                            brinco_eletronico = str(int(brinco_e_raw))
                        elif isinstance(brinco_e_raw, (datetime, pd.Timestamp)):
                            brinco_eletronico = str(brinco_e_raw.date())
                        else:
                            brinco_eletronico = str(brinco_e_raw).strip()
                
                # Data de Nascimento (opcional)
                data_nascimento = None
                if 'Data de Nascimento' in best_mapping:
                    data_n_col = best_mapping['Data de Nascimento']
                    data_n_raw = row[data_n_col]
                    
                    if not pd.isna(data_n_raw):
                        try:
                            if isinstance(data_n_raw, (datetime, pd.Timestamp)):
                                data_nascimento = data_n_raw.date()
                            else:
                                # Tenta converter string para data
                                data_nascimento = pd.to_datetime(data_n_raw).date()
                        except:
                            print(f"Aviso: Não foi possível converter data de nascimento '{data_n_raw}'")
                
                # Data de Entrada (opcional)
                data_entrada = None
                if 'Data de Entrada' in best_mapping:
                    data_e_col = best_mapping['Data de Entrada']
                    data_e_raw = row[data_e_col]
                    
                    if not pd.isna(data_e_raw):
                        try:
                            if isinstance(data_e_raw, (datetime, pd.Timestamp)):
                                data_entrada = data_e_raw.date()
                            else:
                                # Tenta converter string para data
                                data_entrada = pd.to_datetime(data_e_raw).date()
                        except:
                            print(f"Aviso: Não foi possível converter data de entrada '{data_e_raw}'")
                
                # Usar data atual se a data de entrada não estiver definida
                if not data_entrada:
                    data_entrada = timezone.now().date()
                
                # Peso de Entrada (opcional)
                peso_entrada = None
                if 'Peso de Entrada' in best_mapping:
                    peso_col = best_mapping['Peso de Entrada']
                    peso_raw = row[peso_col]
                    
                    if not pd.isna(peso_raw):
                        try:
                            if isinstance(peso_raw, (int, float, np.number)):
                                peso_entrada = Decimal(str(peso_raw))
                            else:
                                # Tenta converter string para decimal
                                peso_entrada = Decimal(str(peso_raw).replace(',', '.'))
                        except:
                            print(f"Aviso: Não foi possível converter peso '{peso_raw}'")
                
                # Valor de Compra (opcional)
                valor_compra = None
                if 'Valor de Compra' in best_mapping:
                    valor_col = best_mapping['Valor de Compra']
                    valor_raw = row[valor_col]
                    
                    if not pd.isna(valor_raw):
                        try:
                            if isinstance(valor_raw, (int, float, np.number)):
                                valor_compra = Decimal(str(valor_raw))
                            else:
                                # Tenta converter string para decimal
                                valor_compra = Decimal(str(valor_raw).replace(',', '.'))
                        except:
                            print(f"Aviso: Não foi possível converter valor '{valor_raw}'")
                
                # ==== CRIAR O ANIMAL ====
                animal = Animal(
                    brinco_visual=brinco_visual,
                    brinco_eletronico=brinco_eletronico,
                    data_nascimento=data_nascimento,
                    data_entrada=data_entrada,
                    peso_entrada=peso_entrada,
                    valor_compra=valor_compra,
                    lote=lote,
                    raca=raca,
                    categoria=categoria,
                    pasto_atual=pasto,
                    usuario=request.user
                )
                
                # Salvar o animal
                animal.save()
                imported_count += 1
                print(f"Animal '{brinco_visual}' importado com sucesso!")
                
            except Exception as e:
                print(f"ERRO ao processar linha {index+1}: {str(e)}")
                import traceback
                traceback.print_exc()
                errors.append(f"Linha {index+1}: Erro - {str(e)}")
        
        # === RETORNO DOS RESULTADOS ===
        
        # Se houver erros, retorna eles junto com a contagem de importados
        if errors:
            return JsonResponse({
                'status': 'partial_success',
                'message': f'Importação parcial: {imported_count} animais importados com {len(errors)} erros',
                'errors': errors
            })
        
        # Se não houver erros, retorna sucesso
        return JsonResponse({
            'status': 'success',
            'message': f'{imported_count} animais importados com sucesso'
        })
        
    except Exception as e:
        # Log detalhado do erro
        print(f"ERRO GERAL: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Em caso de erro na leitura do arquivo
        return JsonResponse({
            'status': 'error',
            'message': f'Erro ao processar arquivo: {str(e)}'
        })
    
    # Se não for POST, retorna para a página de importação
    return render(request, 'animais/animal_import.html', context)

def map_columns(columns):
    """Função para mapear colunas da planilha para os campos do sistema"""
    column_mapping = {}
    print("\n===== MAPEAMENTO DETALHADO DE COLUNAS =====")
    print(f"Colunas para mapear: {list(columns)}")
    
    # Mapeamento extremamente flexível de colunas (mais termos e variações)
    mapping_terms = {
        'Brinco Visual': ['brinco visual', 'brinco', 'visual', 'id', 'identificação', 'identificacao', 
                           'número visual', 'numero visual', 'identificador', 'número', 'numero',
                           'código', 'codigo', 'numeração', 'numeracao', 'brinco_visual', 'id animal'],
        'Brinco Eletrônico': ['brinco eletrônico', 'brinco eletronico', 'eletrônico', 'eletronico', 
                               'chip', 'rfid', 'tag', 'brinco_eletronico', 'eletrônica', 'eletronica'],
        'Nome do Lote': ['nome do lote', 'lote', 'grupo', 'nome lote', 'id lote', 'id do lote', 
                           'identificação do lote', 'identificacao do lote', 'nome_lote', 'lote nome',
                           'grupo de animais', 'grupo animal', 'conjunto', 'nome do grupo'],
        'Data de Nascimento': ['data de nascimento', 'nascimento', 'dt nascimento', 'dt. nascimento', 
                                'data nasc', 'dt nasc', 'data nascimento', 'data_nascimento', 'nascido em',
                                'nasc', 'data de nasc'],
        'Data de Entrada': ['data de entrada', 'entrada', 'dt entrada', 'dt. entrada', 'data entr', 
                             'dt entr', 'data entrada', 'data_entrada', 'dt_entrada', 'data de início',
                             'data de inicio', 'inicio'],
        'Raça': ['raça', 'raca', 'breed', 'tipo', 'tipo de raça', 'tipo de raca', 'tipo raça', 
                  'tipo raca', 'raça animal', 'raca animal', 'genética', 'genetica'],
        'Categoria': ['categoria', 'tipo', 'classe', 'classificação', 'classificacao', 'group', 'grupo', 'tipo animal', 
                       'tipo de animal', 'finalidade', 'funcao', 'função'],
        'Peso de Entrada': ['peso de entrada', 'peso', 'weight', 'peso entrada', 'kg', 'quilos', 
                             'kilos', 'massa', 'peso_entrada', 'peso em kg', 'peso (kg)'],
        'Valor de Compra': ['valor de compra', 'valor', 'preço', 'preco', 'price', 'custo', 
                             'cost', 'valor compra', 'preço de compra', 'preco de compra', 
                             'valor_compra', 'custo de aquisição', 'custo de aquisicao'],
        'Pasto Atual': ['pasto atual', 'pasto', 'paddock', 'área', 'area', 'pasture', 
                          'local atual', 'localização atual', 'localizacao atual', 'pasto_atual',
                          'atual pasto', 'localidade', 'lugar']
    }
    
    # Se temos nomes exatos de colunas, podemos mapear diretamente
    for col in columns:
        col_str = str(col).strip()
        # Caso direto - o nome da coluna é exatamente igual ao nome do campo
        for field in mapping_terms.keys():
            if col_str.lower() == field.lower():
                column_mapping[field] = col
                print(f"Mapeamento direto: '{col}' -> '{field}'")
    
    # Verificar se fizemos todos os mapeamentos
    required_fields = ['Brinco Visual', 'Nome do Lote', 'Raça', 'Categoria', 'Pasto Atual']
    missing_fields = [field for field in required_fields if field not in column_mapping]
    
    if not missing_fields:
        print("Todos os campos obrigatórios encontrados com mapeamento direto!")
        return column_mapping
    
    print(f"Campos ainda não mapeados: {missing_fields}")
    print("Tentando mapeamento por similaridade...")
    
    # Função para calcular similaridade entre strings
    def string_similarity(s1, s2):
        s1 = s1.lower().strip()
        s2 = s2.lower().strip()
        
        # Correspondência exata
        if s1 == s2:
            return 1.0
        
        # Verificar se uma string contém a outra
        if s1 in s2 or s2 in s1:
            return 0.8
        
        # Verificar palavras individuais
        words1 = set(s1.split())
        words2 = set(s2.split())
        common_words = words1.intersection(words2)
        
        if common_words:
            return len(common_words) / max(len(words1), len(words2))
        
        return 0.0
    
    # Para cada coluna na planilha, tentar encontrar a melhor correspondência para campos não mapeados
    columns_to_try = [col for col in columns if not any(col == v for v in column_mapping.values())]
    
    for field in missing_fields:
        print(f"\nTentando encontrar correspondência para '{field}'")
        best_match = None
        best_score = 0.3  # Limiar mínimo de correspondência
        
        for col in columns_to_try:
            col_lower = str(col).lower().strip()
            if not col_lower:  # Pular colunas com nomes vazios
                continue
                
            # Verificar contra todos os termos possíveis para este campo
            for term in mapping_terms[field]:
                score = string_similarity(col_lower, term)
                print(f"  Comparando '{col}' com '{term}' -> score: {score:.2f}")
                if score > best_score:
                    best_score = score
                    best_match = col
        
        if best_match:
            column_mapping[field] = best_match
            print(f"  -> Mapeado '{field}' para coluna '{best_match}' (score: {best_score:.2f})")
        else:
            print(f"  -> Não foi possível encontrar correspondência para '{field}'")
    
    # Tentar mapeamento heurístico para campos que ainda estão faltando
    still_missing = [field for field in required_fields if field not in column_mapping]
    if still_missing and columns_to_try:
        print("\nTentando mapeamento heurístico para campos restantes...")
        
        # Tente fazer correspondência com base na posição típica
        typical_order = ['Brinco Visual', 'Brinco Eletrônico', 'Nome do Lote', 'Raça', 'Categoria', 'Pasto Atual']
        
        # Se temos pelo menos algumas colunas com o mesmo número que nossas colunas típicas
        if len(columns_to_try) >= len(still_missing):
            # Tente atribuir pela posição 
            for field in still_missing:
                try:
                    target_position = typical_order.index(field)
                    if target_position < len(columns_to_try):
                        column_mapping[field] = columns_to_try[target_position]
                        print(f"Mapeamento heurístico: coluna {target_position} '{columns_to_try[target_position]}' -> '{field}'")
                except ValueError:
                    pass
    
    print("Mapeamento final de colunas:", column_mapping)
    return column_mapping

# Função para identificar o índice da linha de cabeçalho
def find_header_row(excel_file):
    """Identifica a linha que contém o cabeçalho na planilha."""
    try:
        print("Iniciando detecção de cabeçalho")
        # Primeiro, tentamos ler sem especificar um cabeçalho
        xls = pd.ExcelFile(excel_file)
        # Lemos sem parsear datas para evitar erros iniciais
        sheet = pd.read_excel(xls, sheet_name=0, header=None)
        
        # Mostrar as primeiras 10 linhas para debug
        print("Primeiras 10 linhas da planilha:")
        for i in range(min(10, len(sheet))):
            row_values = [str(val).strip() if not pd.isna(val) else '' for val in sheet.iloc[i]]
            print(f"Linha {i}: {' | '.join(row_values)}")
        
        # Procuramos por uma linha que tenha as colunas esperadas
        header_keywords = ['brinco visual', 'brinco', 'lote', 'raça', 'raca', 'categoria', 'pasto']
        
        # Verificamos até a linha 15 (mais que suficiente para instruções)
        for i in range(min(15, len(sheet))):
            row = sheet.iloc[i]
            row_values = [str(val).strip().lower() if not pd.isna(val) else '' for val in row]
            row_text = ' '.join(row_values).lower()
            
            # Contar quantas keywords encontramos nesta linha
            matches = sum(1 for keyword in header_keywords if keyword in row_text)
            print(f"Linha {i}: {matches} correspondências de palavras-chave. Texto: {row_text[:100]}...")
            
            # Se pelo menos 2 dos termos chave estiverem na linha, provavelmente é o cabeçalho
            if matches >= 2:
                print(f"Cabeçalho encontrado na linha {i}")
                return i
        
        # Se não encontrou, assume que é a primeira linha
        print("Nenhum cabeçalho óbvio encontrado, usando linha 0")
        return 0
    except Exception as e:
        print(f"Erro ao detectar cabeçalho: {str(e)}")
        traceback.print_exc()
        return 0