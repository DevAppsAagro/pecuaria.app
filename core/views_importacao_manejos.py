import pandas as pd
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from .models import Animal, Pesagem, ManejoSanitario, Lote, Pasto
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import json

@login_required
def manejo_template_download(request):
    """
    Gera um arquivo Excel como modelo para importação de manejos.
    O modelo inclui colunas para pesagens e manejos sanitários.
    """
    # Criar um novo workbook e selecionar a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Importação Manejos"
    
    # Definir estilos
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir cabeçalhos
    headers = [
        "Brinco do Animal*", 
        "Data do Manejo*",
        "Peso (kg)",
        "Fazer Manejo Sanitário (S/N)",
        "Tipo de Manejo",
        "Insumo",
        "Dias para Próximo Manejo",
        "Observação",
        "Fazer Apartação (S/N)",
        "Peso de Referência",
        "Lote Acima",
        "Pasto Acima",
        "Lote Abaixo",
        "Pasto Abaixo"
    ]
    
    # Adicionar cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Adicionar informações de ajuda na segunda linha (opcional)
    help_texts = [
        "Brinco visual ou eletrônico",
        "Formato: DD/MM/AAAA",
        "Usar ponto como separador decimal",
        "S para sim, N para não",
        "Ex: Vacinação, Vermifugação",
        "Nome do insumo utilizado",
        "Número inteiro",
        "Texto livre",
        "S para sim, N para não",
        "Peso para apartação (kg)",
        "Nome do lote para animais acima do peso",
        "Nome do pasto para animais acima do peso",
        "Nome do lote para animais abaixo do peso",
        "Nome do pasto para animais abaixo do peso"
    ]
    
    for col_num, help_text in enumerate(help_texts, 1):
        cell = ws.cell(row=2, column=col_num, value=help_text)
        cell.font = Font(name='Arial', size=10, italic=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Adicionar exemplo na terceira linha
    example = [
        "B001",
        datetime.datetime.now().strftime("%d/%m/%Y"),
        "450.5",
        "S",
        "Vacinação",
        "Vacina Febre Aftosa",
        "180",
        "Aplicação na tábua do pescoço",
        "S",
        "400",
        "Lote Engorda",
        "Pasto A",
        "Lote Recria",
        "Pasto B"
    ]
    
    for col_num, ex_value in enumerate(example, 1):
        cell = ws.cell(row=3, column=col_num, value=ex_value)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Ajustar largura das colunas
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        column_width = max(len(header), len(help_texts[col_num-1]), len(str(example[col_num-1]))) + 2
        ws.column_dimensions[column_letter].width = column_width
    
    # Congelar a primeira linha
    ws.freeze_panes = 'A4'
    
    # Preparar a resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=modelo_importacao_manejos.xlsx'
    
    # Salvar o workbook para a resposta
    wb.save(response)
    
    return response

@login_required
@require_http_methods(["POST"])
def manejo_import(request):
    """
    Processa a importação de manejos a partir de um arquivo Excel.
    Cria registros nas tabelas de Pesagem e ManejoSanitario conforme necessário.
    """
    try:
        # Verificar se um arquivo foi enviado
        if 'import_file' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'Nenhum arquivo enviado'})
        
        import_file = request.FILES['import_file']
        skip_header = request.POST.get('skip_header', 'on') == 'on'
        
        # Determinar o formato do arquivo
        file_ext = import_file.name.split('.')[-1].lower()
        
        # Ler o arquivo com pandas
        if file_ext == 'csv':
            df = pd.read_csv(import_file, encoding='utf-8-sig')
        elif file_ext in ['xlsx', 'xls']:
            df = pd.read_excel(import_file)
        else:
            return JsonResponse({'success': False, 'message': 'Formato de arquivo não suportado. Use .xlsx, .xls ou .csv'})
        
        # Pular o cabeçalho se necessário
        if skip_header and len(df) > 0:
            df = df.iloc[1:]
        
        # Renomear colunas para facilitar o acesso
        df.columns = [
            "brinco", "data", "peso", "fazer_manejo", "tipo_manejo", 
            "insumo", "dias_proximo", "observacao", "fazer_apartacao",
            "peso_referencia", "lote_acima", "pasto_acima", "lote_abaixo", "pasto_abaixo"
        ]
        
        # Inicializar contadores
        total_processados = 0
        pesagens_criadas = 0
        manejos_criados = 0
        apartacoes_realizadas = 0
        erros = []
        
        # Processar cada linha do DataFrame
        for index, row in df.iterrows():
            try:
                # Verificar se o brinco existe
                brinco = str(row['brinco']).strip()
                if not brinco:
                    erros.append(f"Linha {index+2}: Brinco não informado")
                    continue
                
                # Buscar o animal pelo brinco
                animal = Animal.objects.filter(Q(brinco_visual=brinco) | Q(brinco_eletronico=brinco)).first()
                if not animal:
                    erros.append(f"Linha {index+2}: Animal com brinco {brinco} não encontrado")
                    continue
                
                # Processar a data
                try:
                    if isinstance(row['data'], str):
                        # Tentar diferentes formatos de data
                        try:
                            data_manejo = datetime.datetime.strptime(row['data'], "%d/%m/%Y").date()
                        except ValueError:
                            try:
                                data_manejo = datetime.datetime.strptime(row['data'], "%Y-%m-%d").date()
                            except ValueError:
                                data_manejo = datetime.datetime.strptime(row['data'], "%d-%m-%Y").date()
                    else:
                        # Assumir que é um objeto datetime do pandas
                        data_manejo = row['data'].date() if hasattr(row['data'], 'date') else row['data']
                except Exception as e:
                    erros.append(f"Linha {index+2}: Erro ao processar data: {str(e)}")
                    continue
                
                # Registrar pesagem se o peso for informado
                peso = None
                if not pd.isna(row['peso']):
                    try:
                        peso = float(str(row['peso']).replace(',', '.'))
                        pesagem = Pesagem(
                            animal=animal,
                            peso=peso,
                            data=data_manejo,
                            usuario=request.user
                        )
                        pesagem.save()
                        pesagens_criadas += 1
                    except Exception as e:
                        erros.append(f"Linha {index+2}: Erro ao registrar pesagem: {str(e)}")
                
                # Registrar manejo sanitário se solicitado
                fazer_manejo = str(row['fazer_manejo']).strip().upper() in ['S', 'SIM', 'Y', 'YES', 'TRUE', '1']
                if fazer_manejo:
                    try:
                        tipo_manejo = str(row['tipo_manejo']).strip()
                        insumo = str(row['insumo']).strip() if not pd.isna(row['insumo']) else ''
                        
                        # Converter dias_proximo para inteiro, tratando valores nulos
                        dias_proximo = 0
                        if not pd.isna(row['dias_proximo']):
                            try:
                                dias_proximo = int(float(str(row['dias_proximo']).replace(',', '.')))
                            except:
                                dias_proximo = 0
                        
                        observacao = str(row['observacao']).strip() if not pd.isna(row['observacao']) else ''
                        
                        manejo = ManejoSanitario(
                            animal=animal,
                            data=data_manejo,
                            insumo=insumo,
                            tipo_manejo=tipo_manejo,
                            dias_proximo_manejo=dias_proximo,
                            observacao=observacao,
                            usuario=request.user
                        )
                        manejo.save()
                        manejos_criados += 1
                    except Exception as e:
                        erros.append(f"Linha {index+2}: Erro ao registrar manejo sanitário: {str(e)}")
                
                # Realizar apartação se solicitado e se o peso foi informado
                fazer_apartacao = str(row['fazer_apartacao']).strip().upper() in ['S', 'SIM', 'Y', 'YES', 'TRUE', '1']
                if fazer_apartacao and peso is not None:
                    try:
                        # Obter o peso de referência
                        peso_referencia = 0
                        if not pd.isna(row['peso_referencia']):
                            peso_referencia = float(str(row['peso_referencia']).replace(',', '.'))
                        
                        # Determinar lote e pasto com base no peso
                        if peso > peso_referencia:
                            # Animal acima do peso de referência
                            lote_nome = str(row['lote_acima']).strip() if not pd.isna(row['lote_acima']) else None
                            pasto_nome = str(row['pasto_acima']).strip() if not pd.isna(row['pasto_acima']) else None
                        else:
                            # Animal abaixo do peso de referência
                            lote_nome = str(row['lote_abaixo']).strip() if not pd.isna(row['lote_abaixo']) else None
                            pasto_nome = str(row['pasto_abaixo']).strip() if not pd.isna(row['pasto_abaixo']) else None
                        
                        # Buscar lote e pasto pelo nome
                        lote = None
                        pasto = None
                        
                        if lote_nome:
                            lote = Lote.objects.filter(nome__iexact=lote_nome, fazenda=animal.fazenda_atual).first()
                            if not lote:
                                erros.append(f"Linha {index+2}: Lote '{lote_nome}' não encontrado na fazenda atual do animal")
                        
                        if pasto_nome:
                            pasto = Pasto.objects.filter(nome__iexact=pasto_nome, fazenda=animal.fazenda_atual).first()
                            if not pasto:
                                erros.append(f"Linha {index+2}: Pasto '{pasto_nome}' não encontrado na fazenda atual do animal")
                        
                        # Atualizar animal com novo lote e pasto
                        if lote:
                            animal.lote = lote
                        
                        if pasto:
                            animal.pasto_atual = pasto
                        
                        if lote or pasto:
                            animal.save()
                            apartacoes_realizadas += 1
                    except Exception as e:
                        erros.append(f"Linha {index+2}: Erro ao realizar apartação: {str(e)}")
                
                total_processados += 1
                
            except Exception as e:
                erros.append(f"Linha {index+2}: Erro ao processar: {str(e)}")
        
        # Preparar resposta
        resultado = {
            'success': True,
            'total_processados': total_processados,
            'pesagens_criadas': pesagens_criadas,
            'manejos_criados': manejos_criados,
            'apartacoes_realizadas': apartacoes_realizadas,
            'erros': erros,
            'message': f"Importação concluída: {pesagens_criadas} pesagens, {manejos_criados} manejos sanitários e {apartacoes_realizadas} apartações realizadas."
        }
        
        if erros:
            resultado['message'] += f" Ocorreram {len(erros)} erros durante a importação."
        
        return JsonResponse(resultado)
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False, 
            'message': f'Erro ao processar arquivo: {str(e)}',
            'traceback': traceback.format_exc()
        })
