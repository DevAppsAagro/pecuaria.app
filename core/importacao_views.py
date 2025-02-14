from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils.timezone import now
from .models import CategoriaAnimal, Lote

def criar_planilha_modelo():
    """
    Cria uma planilha modelo para importação de animais com as colunas necessárias
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Importação de Animais"

    # Define o cabeçalho
    headers = [
        'Brinco Visual*',
        'Brinco Eletrônico',
        'Raça*',
        'Categoria*',
        'Data de Nascimento* (DD/MM/AAAA)',
        'Data de Entrada* (DD/MM/AAAA)',
        'Peso de Entrada (kg)*',
        'Valor de Compra (R$)*',
        'Nome do Lote*',
        'Pasto Atual*'
    ]

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Aplica o cabeçalho e o estilo
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Ajusta a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Adiciona uma linha de exemplo
    example_data = [
        'BR123',
        'E123456',
        'Nelore',
        'Bezerro',
        '01/01/2024',
        '01/01/2024',
        '180.5',
        '2000.00',
        'Lote Engorda 1',
        'Pasto A'
    ]
    
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.alignment = Alignment(horizontal='center')

    # Adiciona uma linha de observação
    ws.cell(row=4, column=1, value='* Campos obrigatórios')
    ws.cell(row=4, column=1).font = Font(bold=True, color="FF0000")

    # Adiciona instruções de preenchimento
    instructions = [
        '- Preencha as datas no formato DD/MM/AAAA (exemplo: 31/12/2024)',
        '- Use ponto (.) como separador decimal para peso e valor',
        '- Use o nome do lote exatamente como aparece no sistema',
        '- Categorias disponíveis: ' + ', '.join([cat.nome for cat in CategoriaAnimal.objects.all()]),
        '- Todos os campos com * são obrigatórios'
    ]
    
    for i, instruction in enumerate(instructions, 5):
        cell = ws.cell(row=i, column=1, value=instruction)
        cell.font = Font(italic=True)
        # Mescla as células para a instrução ocupar várias colunas
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(headers))
        cell.alignment = Alignment(horizontal='left')

    return wb

@login_required
def download_planilha_modelo(request):
    """
    View para download da planilha modelo de importação de animais
    """
    wb = criar_planilha_modelo()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=importacao_animais_modelo.xlsx'
    wb.save(response)
    return response
