import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def criar_planilha_modelo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importação de Animais"

    # Definir cabeçalhos
    headers = [
        'Brinco Visual*',
        'Brinco Eletrônico',
        'ID do Lote*',
        'Data de Nascimento* (DD/MM/AAAA)',
        'Data de Entrada* (DD/MM/AAAA)',
        'Raça*',
        'Categoria*',
        'Peso de Entrada (kg)',
        'Valor de Compra (R$)'
    ]

    # Estilo para cabeçalhos
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    header_font = Font(bold=True)
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar largura da coluna
        ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 2, 15)

    # Adicionar exemplo de dados
    example_data = [
        'BR0001',           # Brinco Visual
        '900001234567890',  # Brinco Eletrônico
        'L001',            # ID do Lote
        '01/01/2023',      # Data de Nascimento
        '15/01/2023',      # Data de Entrada
        'Nelore',          # Raça
        'Bezerro',         # Categoria
        '180.5',           # Peso de Entrada
        '2500.00'          # Valor de Compra
    ]

    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.alignment = Alignment(horizontal='center')

    # Adicionar instruções
    ws2 = wb.create_sheet(title="Instruções")
    instructions = [
        ["Instruções para Preenchimento da Planilha de Importação de Animais"],
        [""],
        ["Campos Obrigatórios (marcados com *)"],
        ["- Brinco Visual: Identificador único do animal"],
        ["- ID do Lote: Código do lote onde o animal será cadastrado"],
        ["- Data de Nascimento: Formato DD/MM/AAAA"],
        ["- Data de Entrada: Formato DD/MM/AAAA"],
        ["- Raça: Nome da raça cadastrada no sistema"],
        ["- Categoria: Nome da categoria cadastrada no sistema"],
        [""],
        ["Campos Opcionais"],
        ["- Brinco Eletrônico: Identificador eletrônico único"],
        ["- Peso de Entrada: Peso em kg (use ponto para decimais)"],
        ["- Valor de Compra: Valor em R$ (use ponto para decimais)"],
        [""],
        ["Observações Importantes"],
        ["1. Não altere a ordem ou nome das colunas"],
        ["2. Certifique-se que o lote informado existe no sistema"],
        ["3. Use o exemplo fornecido como referência para o formato dos dados"],
        ["4. Raças e Categorias devem estar previamente cadastradas no sistema"]
    ]

    for row_num, instruction in enumerate(instructions, 1):
        cell = ws2.cell(row=row_num, column=1, value=instruction[0])
        if row_num == 1:
            cell.font = Font(bold=True, size=12)
        
    # Ajustar largura da coluna de instruções
    ws2.column_dimensions['A'].width = 70

    return wb

if __name__ == '__main__':
    wb = criar_planilha_modelo()
    wb.save('importacao_animais_modelo.xlsx')
